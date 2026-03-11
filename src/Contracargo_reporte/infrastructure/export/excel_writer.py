from pathlib import Path
from typing import Any
import logging

import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import date, datetime


LOGGER = logging.getLogger(__name__)


def exportar_pestana_texto(
    rows: list[tuple[Any, ...]],
    cols: list[str],
    ruta: Path,
    sheet_name: str,
    freeze_panes: str | None = "A2",
    date_columns: set[str] | None = None,
    date_format: str = "dd/mm/yyyy",
) -> None:
    """
    Escribe una hoja en un Excel existente sin eliminar la pestaña:
    - Si existe: limpia celdas y reescribe.
    - Si no existe: crea la hoja.
    """
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    if ruta.exists():
        wb = load_workbook(ruta)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _clear_sheet(ws)
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Si es un libro nuevo, elimina la hoja por defecto si queda vacia.
    if (
        len(wb.sheetnames) > 1
        and "Sheet" in wb.sheetnames
        and wb["Sheet"].max_row == 1
        and wb["Sheet"].max_column == 1
    ):
        wb.remove(wb["Sheet"])

    date_columns = date_columns or set()
    date_col_indices = {idx + 1 for idx, name in enumerate(cols) if name in date_columns}

    for col_idx, header in enumerate(cols, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in date_col_indices and isinstance(value, (date, datetime)):
                cell.number_format = date_format

    if freeze_panes:
        ws.freeze_panes = freeze_panes
    wb.save(ruta)


def _clear_sheet(ws) -> None:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None
